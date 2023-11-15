import json
import logging
import os
from tqdm import tqdm
from datetime import datetime, timedelta
import shutil

import ckanapi
from openpyxl.reader.excel import load_workbook

log = logging.getLogger(__name__)
log.setLevel(logging.INFO)

CONFIG_FILENAME = os.getenv('CONFIG_FILENAME', 'config.json')
root_dir = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(root_dir, CONFIG_FILENAME)

with open(CONFIG_PATH, 'r') as config_file:
    CONFIG = json.loads(config_file.read())['config']

RESOURCE_FOLDER = os.path.join(root_dir, CONFIG['resources_folders'])
DATASETS_FILE = root_dir + CONFIG['datasets_file']

MONTHS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December"
}


def mutable_dataset_dict(title, name, month, resources, **kwargs):
    return {
        "title": title,
        "name": name,
        "type": "family-medicine",
        "notes": kwargs.get("notes", ""),
        "owner_org": "who-romania",
        "maintainer": kwargs.get("maintainer", "Georgiana Afumateanu"),
        "maintainer_email": kwargs.get("maintainer_email", "afumateanug@who.int"),
        "month": month,
        "groups": [{"name": "family-medicine"}],
        "tags": [{"name": "Data"}],
        "year": kwargs.get("year", 2023),
        "resources": resources
    }


def mutable_resource_dict(name, file_path, week, family_doctor):
    return {
        "name": name,
        "filename": file_path,
        "format": "XLSX",
        "week": week,
        "family_doctor": family_doctor
    }


def traverse_folder_tree():
    """
    Traverses a folder tree and returns a list of all file paths in the tree.
    
    Returns:
        dict: A list of all file paths in the folder tree.
    """
    subfolders = [int(dir_name) for dir_name in os.listdir(CONFIG['data_folder'])]
    subfolders.sort(reverse=True)
    files = []
    for subfolder in subfolders:
        for dirpath, dirnames, filenames in os.walk(CONFIG['data_folder'] + '/' + str(subfolder)):
            year = '20' + (str(subfolder)[:2])
            month = (str(subfolder).split('23')[1])
            folders_dict = {"year": year, "month": month, "files": []}
            for file in filenames:
                if "month" in file.lower():  # skip monthly files
                    continue
                if os.path.splitext(file)[1] in ['.xls', '.xlsx']:
                    folders_dict["files"].append(os.path.join(dirpath, file))
            files.append(folders_dict)
    return files


def read_resource_sheet(filename, workbook, sheet):
    """
    Reads data from an Excel sheet, creates a new folder, copies a template Excel file to the new folder,
    populates the template with data from the sheet, and returns a dictionary containing information about the new resource.
    """
    active = workbook[sheet]
    report_date = active['B2'].value
    fd_name = sheet.split('FD ')[1]

    if not isinstance(report_date, datetime):
        log.warning(f"Report date {report_date} is not a date.")
        log.warning("Failed to use report date, attempting to load from report period ...")
        with open(root_dir + '/logs/report_date_error.csv', 'a') as file:
            file.write(f"{filename},{sheet},{report_date}\n")
        report_date = active['B3'].value
        if not isinstance(report_date, datetime):
            log.error(f"Report date {report_date} is not a date.")

    if report_date.weekday() != 4:
        log.warning(f"Report date {report_date} is not a Friday.")
        with open(root_dir + '/logs/not_friday_error.csv', 'a') as file:
            file.write(f"{filename},{sheet},{report_date},{report_date.strftime('%A')}\n")

    year = str(report_date.year % 100)
    month = "{:02}".format(report_date.month)
    day = "{:02}".format(report_date.day)
    week = year + '-' + month + '-' + day
    week_number = str(report_date.isocalendar()[1])

    new_folder_path = os.path.join(root_dir, "resources/family-medicine-reports", month, week, sheet)

    # Loading the template file
    template_path = os.path.join(root_dir, CONFIG['template_file'])
    template_workbook = load_workbook(template_path)
    template_sheet = template_workbook.active

    start_row, end_row = 3, 35
    start_col, end_col = 4, 23  # A to W

    sheet_contains_data = False
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell_value = active.cell(row=row+5, column=col).value
            if cell_value is not None:
                sheet_contains_data = True
            template_sheet.cell(row=row, column=col, value=cell_value)

    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)

    save_path = new_folder_path + '/report.xlsx'

    if sheet_contains_data:
        template_workbook.save(save_path)
    else:
        # Copy instead, because using pyexcel corrupts pandas reading
        shutil.copy2(template_path, new_folder_path)
        base_name = os.path.basename(template_path)
        copied_filename = os.path.join(new_folder_path, base_name)
        os.rename(copied_filename, save_path)

    return mutable_resource_dict("Report from Family Doctor " + fd_name + " for week number " + week_number,
                                 'report.xlsx',
                                 week,
                                 fd_name)

def get_dates_of_weekday_in_month(month, weekday=4, format="%d %b %Y"):
    month_start = datetime.strptime(month, "%Y-%m")
    date = month_start + timedelta((weekday-month_start.weekday())%7)
    week = timedelta(weeks=1)
    dates = []
    while date.month == month_start.month:
        dates.append(date.strftime(format))
        date = date + week
    return dates

def load_file_sheets(file):
    log.info("Reading file: ", file)
    workbook = load_workbook(file)
    sheets = [item for item in workbook.sheetnames if item.startswith('FD')]
    sheets.sort()
    all_resources = []
    for sheet in sheets:
        all_resources.append(read_resource_sheet(file, workbook, sheet))
    return all_resources


def generate_dataset_dict():
    dataset_dict = {
        "datasets": []
    }
    fd_files = traverse_folder_tree()
    for file_dict in tqdm(fd_files):
        resources = []
        for file in file_dict["files"]:
            resources.extend(load_file_sheets(file))

        dataset_dict["datasets"].append(mutable_dataset_dict(
            title="Family Medicine Data",
            name="family-medicine-data" + "-" + (file_dict["month"]) + "-" + (file_dict["year"]),
            month=file_dict["year"] + "-" + file_dict["month"],
            year=file_dict["year"],
            notes=(f"Containing weekly reports to WHO Romania country office from selected family medicine "
                   f"doctors for the month of {MONTHS[int(file_dict['month'])]}. This data covers total number "
                   f"of consultation, as well as purpose of the consultation, syndromic surveillance, referrals "
                   f"and vaccination. "),
            resources=resources
        ))

    with open(root_dir + '/resources/datasets.json', 'w') as json_file:
        json.dump(dataset_dict, json_file, indent=4)


def load_datasets(ckan):
    """
    Helper method to load datasets from the DATASETS_FILE config file
    :param ckan: ckanapi instance
    :return: None
    """
    with open(DATASETS_FILE, 'r') as datasets_file:
        datasets = json.load(datasets_file)['datasets']
        for dataset in tqdm(datasets):
            resources = dataset.pop('resources', [])
            dataset['resources'] = []
            try:
                dataset = ckan.action.package_create(**dataset)
                log.info(f"Created dataset {dataset['name']}")
            except ckanapi.errors.ValidationError:
                try:
                    log.warning(f"Dataset {dataset['name']} might exist. Will try to update.")
                    id = ckan.action.package_show(id=dataset['name'])['id']
                    dataset = ckan.action.package_update(id=id, **dataset)
                    log.info(f"Updated dataset {dataset['name']}")
                except ckanapi.errors.ValidationError as e:
                    log.error(f"Can't create dataset {dataset['name']}: {e.error_dict}")
                    continue
            for resource in resources:
                # get the month from the resource week
                month = resource['week'].split('-')[1]
                file_path = os.path.join(RESOURCE_FOLDER, month, resource["week"], "FD " + resource["family_doctor"],
                                         resource['filename'])
                resource['package_id'] = dataset['id']
                try:

                    try:
                        with open(file_path, 'rb') as res_file:
                            resource = ckan.call_action(
                                'resource_create',
                                resource,
                                files={'upload': res_file}
                            )
                    except Exception:
                        log.warning(f"Failed to load {resource['name']} for {dataset['name']}.  Trying again...")
                        with open(file_path, 'rb') as res_file:
                            resource = ckan.call_action(
                                'resource_create',
                                resource,
                                files={'upload': res_file}
                            )
                    log.info(f"Created resource {resource['name']}")
                except ckanapi.errors.ValidationError as e:
                    log.error(f"Can't create resource {resource['name']}: {e.error_dict}")


if __name__ == '__main__':
    ckan = ckanapi.RemoteCKAN(CONFIG['ckan_url'], apikey=CONFIG['ckan_api_key'])
    generate_dataset_dict()
    if not os.path.exists(root_dir + '/resources/datasets.json'):
        log.error("Failed to generate datasets.json")
        exit(1)
    load_datasets(ckan)
